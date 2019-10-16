# 這個是check.py 的改寫

from pathlib import Path
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import time

#（1）讀取excell 資料的函數

def fileread(name,i,j,k):

    X =datetime.today() # 再檔案名稱加入當天日期
    cusy = X.year
    cusm = X.month
    cusd = X.day
    idname = str(cusy)+'-'+str(cusm)+'-'+str(cusd)

    #確認資料是否存在

    filecheck = Path(name+".xlsx")

    if not filecheck.exists():
        print('Oops, '+name +'doesnt exist!')
    #讀取舊的資料表設定
    else:
        wb = load_workbook(filename = name +'.xlsx')
        ws = wb.active
        mxrow = ws.max_row #求列的最大數

        #寫入新的資料表設定
        wbr = Workbook()
        ws2 = wbr.create_sheet(title=name)
        dest_filename = name+idname+'.xlsx'

        for row in range(0,mxrow):
            ws2.cell(row=row+1,column=1).value = name
            ws2.cell(row=row+1,column=2).value = ws.cell(row=row+1,column=1).value
            ws2.cell(row=row+1,column=3).value = ws.cell(row=row+1,column=4).value
            ws2.cell(row=row+1,column=4).value = ws.cell(row=row+1,column=6).value
        #讀取同步轉存到其他資料表
            
        wbr.save(filename = dest_filename)
       


# (2）使用函數

fileread('代收-一銀文',1,4,7)
fileread('代收-臺企公司',1,4,5)
fileread('代收-臺企榮和',1,4,5)
fileread('代收-臺企鑫',1,4,5)
fileread('代收-聯邦茹',2,3,4)

