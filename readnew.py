# 這個是測試存檔資料檔案- 可以指定對應的excell 欄位

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = load_workbook(filename = '代收-一銀文.xlsx')
ws = wb.active
mxrow = ws.max_row
print(mxrow)

wbr = Workbook()
ws2 = wbr.create_sheet(title="如如測試")
dest_filename = '銀行測試.xlsx'

ws2['C4'] ='TEST VALUE'

for row in range(5,mxrow):
		ws2.cell(row=row+1,column=1).value =ws.cell(row=row+1,column=1).value
		ws2.cell(row=row+1,column=2).value =ws.cell(row=row+1,column=4).value		
	

wbr.save(filename = dest_filename)


